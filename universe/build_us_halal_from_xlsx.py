"""
Extracts Ahmed's authoritative US halal universe from
universe/US stocks halal.xlsx and writes:
  - universe/us_halal_full.txt   (all 1,621 tickers, by market cap desc)
  - universe/us_halal_top30.txt  (top 30 — replaces prior HLAL proxy)
  - universe/us_halal_meta.json  (ticker -> {name, isin, sector, mcap})

Ticker format in Ahmed's file: "NVDA US" — strip " US" suffix for yfinance.
"""

from __future__ import annotations
import json
import os
from openpyxl import load_workbook

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(THIS_DIR, "US stocks halal.xlsx")
FULL = os.path.join(THIS_DIR, "us_halal_full.txt")
TOP30 = os.path.join(THIS_DIR, "us_halal_top30.txt")
META = os.path.join(THIS_DIR, "us_halal_meta.json")


def clean_ticker(raw: str) -> str:
    """Normalise Ahmed's '<TICKER> <COUNTRY>' format to yfinance convention.

    Rules:
      - Take the first whitespace-separated segment (drops ' US', ' TW', ' NL',
        ' CH', ' DE', ' GB', ' IE', etc.).
      - Strip trailing asterisks/footnote markers (SCCO* -> SCCO).
      - Convert dots to dashes for class shares (BRK.B -> BRK-B).
      - Uppercase.
    """
    t = (raw or "").strip()
    if not t:
        return ""
    # first segment before any whitespace = base ticker (drops country code)
    t = t.split()[0]
    # strip trailing non-alphanumeric except '-' (which is class-share marker)
    while t and not (t[-1].isalnum() or t[-1] == "-"):
        t = t[:-1]
    # class-share dot -> dash
    t = t.replace(".", "-")
    return t.upper()


def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if r and r[2]]

    parsed = []
    seen = set()
    dup_count = 0
    for name, isin, ticker_raw, sector, mcap in rows:
        tk = clean_ticker(ticker_raw)
        if not tk:
            continue
        if tk in seen:
            dup_count += 1
            continue
        seen.add(tk)
        parsed.append({"ticker": tk, "raw": ticker_raw, "name": name,
                       "isin": isin, "sector": sector, "market_cap": mcap})

    parsed.sort(key=lambda r: r["market_cap"] or 0, reverse=True)
    print(f"Parsed {len(parsed)} unique tickers from xlsx "
          f"({dup_count} duplicates dropped)")

    with open(FULL, "w", encoding="utf-8") as fh:
        fh.write("# US ADIB halal universe (authoritative, from Ahmed's "
                 "'US stocks halal.xlsx' 2026-05-20)\n")
        fh.write(f"# {len(parsed)} tickers, sorted by market cap desc.\n")
        fh.write("# yfinance convention: bare symbol (e.g. NVDA). "
                 "Class shares use - not . (e.g. BRK-B).\n\n")
        for r in parsed:
            fh.write(f"{r['ticker']}\n")

    with open(TOP30, "w", encoding="utf-8") as fh:
        fh.write("# Top 30 US halal tickers by market cap "
                 "(from Ahmed's authoritative list, 2026-05-20).\n")
        fh.write("# Replaces the HLAL proxy used in Session 1.\n\n")
        for r in parsed[:30]:
            fh.write(f"{r['ticker']}\n")

    with open(META, "w", encoding="utf-8") as fh:
        json.dump({r["ticker"]: {"name": r["name"], "isin": r["isin"],
                                 "sector": r["sector"],
                                 "market_cap": r["market_cap"]}
                   for r in parsed}, fh, indent=2)

    print(f"Wrote {FULL}")
    print(f"Wrote {TOP30}")
    print(f"Wrote {META}")
    print("Top 30:", [r["ticker"] for r in parsed[:30]])


if __name__ == "__main__":
    main()
