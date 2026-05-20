"""
Builds universe/uae_stocks.xlsx — a CEO-curated working list of UAE
Shariah-compliant equities pending Ahmed's authoritative ADIB / DFM list.

Output:
  universe/uae_stocks.xlsx   — three sheets:
      "Halal Universe"   — active working list (loaded by the engine)
      "Excluded"         — known-excluded names (conventional bank/insurance/etc)
      "About"            — methodology + linkage to engine
  universe/uae_tickers.txt   — derived plain-text list, one ticker per line
                               (matches engine universe loader format)

Methodology: AAOIFI sector screen applied to publicly-known ADX + DFM listings.
This is NOT a substitute for Ahmed's authoritative ADIB universe. Drop the
real list at `universe/uae_tickers.txt` to override.
"""

from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(THIS_DIR, "uae_stocks.xlsx")
TXT_PATH = os.path.join(THIS_DIR, "uae_tickers.txt")


# (ticker, name, exchange, sector, cap_tier, halal_status, notes)
# halal_status: HALAL (AAOIFI sector + ratios consistent) / VERIFY (sector
# borderline — Ahmed to confirm) / EXCLUDE_CONV (conventional finance)
HALAL = [
    # --- ADX (Abu Dhabi Securities Exchange) ---
    ("ADIB.AD",       "Abu Dhabi Islamic Bank",          "ADX", "Banking-Islamic",   "Large", "HALAL",  ""),
    ("ADNOCDIST.AD",  "ADNOC Distribution",              "ADX", "Energy-Retail",     "Large", "HALAL",  ""),
    ("ADNOCDRILL.AD", "ADNOC Drilling",                  "ADX", "Energy-Services",   "Large", "HALAL",  ""),
    ("ADNOCGAS.AD",   "ADNOC Gas",                       "ADX", "Energy-Gas",        "Mega",  "HALAL",  ""),
    ("ADNOCLS.AD",    "ADNOC Logistics & Services",      "ADX", "Industrials-Logistics","Mid","HALAL",  ""),
    ("ADPORTS.AD",    "AD Ports Group",                  "ADX", "Industrials-Logistics","Large","HALAL", ""),
    ("AGTHIA.AD",     "Agthia Group",                    "ADX", "Consumer-Staples",  "Mid",   "HALAL",  ""),
    ("ALDAR.AD",      "Aldar Properties",                "ADX", "Real-Estate",       "Large", "HALAL",  ""),
    ("ALPHADHABI.AD", "Alpha Dhabi Holding",             "ADX", "Diversified",       "Large", "HALAL",  ""),
    ("ASMAK.AD",      "Int'l Fish Farming Co (Asmak)",   "ADX", "Consumer-Staples",  "Small", "HALAL",  ""),
    ("BOROUGE.AD",    "Borouge",                         "ADX", "Materials-Petrochem","Large","HALAL",  ""),
    ("DANA.AD",       "Dana Gas",                        "ADX", "Energy-Gas",        "Mid",   "HALAL",  "Halal per AAOIFI; excluded from R4 strategy on perf basis"),
    ("EAND.AD",       "e& (Etisalat)",                   "ADX", "Telecom",           "Mega",  "HALAL",  ""),
    ("EMSTEEL.AD",    "Emirates Steel Arkan",            "ADX", "Materials-Steel",   "Mid",   "HALAL",  ""),
    ("FERTIGLB.AD",   "Fertiglobe",                      "ADX", "Materials-Chemicals","Mid",  "HALAL",  ""),
    ("GULFCEMENT.AD", "Gulf Cement",                     "ADX", "Materials-Cement",  "Small", "HALAL",  ""),
    ("IHC.AD",        "International Holding Company",   "ADX", "Diversified",       "Mega",  "HALAL",  ""),
    ("MULTIPLY.AD",   "Multiply Group",                  "ADX", "Diversified",       "Mid",   "HALAL",  ""),
    ("NMDC.AD",       "National Marine Dredging Co",     "ADX", "Industrials",       "Mid",   "HALAL",  ""),
    ("PRESIGHT.AD",   "Presight AI",                     "ADX", "Technology",        "Mid",   "HALAL",  ""),
    ("PUREHEALTH.AD", "Pure Health",                     "ADX", "Healthcare",        "Large", "HALAL",  ""),
    ("RAKCEC.AD",     "RAK Cement",                      "ADX", "Materials-Cement",  "Small", "HALAL",  ""),
    ("RAKWCT.AD",     "RAK White Cement",                "ADX", "Materials-Cement",  "Small", "HALAL",  ""),
    ("RAPCO.AD",      "RAK Poultry & Feeding",           "ADX", "Consumer-Staples",  "Small", "HALAL",  ""),
    ("SHAREK.AD",     "Sharjah Cement",                  "ADX", "Materials-Cement",  "Small", "HALAL",  ""),
    ("SPACE42.AD",    "Space42 (formerly Yahsat)",       "ADX", "Telecom-Satellite", "Mid",   "HALAL",  ""),
    ("TAQA.AD",       "TAQA (AD National Energy)",       "ADX", "Utilities",         "Mega",  "HALAL",  ""),
    ("UMC.AD",        "Union Cement",                    "ADX", "Materials-Cement",  "Small", "HALAL",  ""),
    ("UMORTH.AD",     "Umm Al Qaiwain General Invest",   "ADX", "Diversified",       "Small", "VERIFY", "Investment co — debt-ratio check needed"),
    ("WAHA.AD",       "Waha Capital",                    "ADX", "Financial-Islamic", "Mid",   "VERIFY", "Islamic-compliant investments — confirm latest ratios"),

    # --- DFM (Dubai Financial Market) ---
    ("AIRARABIA.DU",  "Air Arabia",                      "DFM", "Transport-Airlines","Mid",   "HALAL",  ""),
    ("AMANAT.DU",     "Amanat Holdings",                 "DFM", "Healthcare/Edu",    "Mid",   "HALAL",  ""),
    ("AMLAK.DU",      "Amlak Finance",                   "DFM", "Real-Estate-Islamic","Small","HALAL",  "Islamic real estate finance"),
    ("ARMX.DU",       "Aramex",                          "DFM", "Logistics",         "Mid",   "HALAL",  ""),
    ("DEWA.DU",       "DEWA",                            "DFM", "Utilities",         "Mega",  "HALAL",  ""),
    ("DEYAAR.DU",     "Deyaar Development",              "DFM", "Real-Estate",       "Small", "HALAL",  ""),
    ("DFM.DU",        "Dubai Financial Market",          "DFM", "Financial-Exchange","Mid",   "VERIFY", "Exchange operator — revenue mix check"),
    ("DIB.DU",        "Dubai Islamic Bank",              "DFM", "Banking-Islamic",   "Large", "HALAL",  ""),
    ("DIC.DU",        "Dubai Islamic Ins. & Reinsurance","DFM", "Insurance-Takaful", "Small", "HALAL",  ""),
    ("DTC.DU",        "Dubai Taxi Company",              "DFM", "Transport",         "Mid",   "HALAL",  ""),
    ("DUBAIINV.DU",   "Dubai Investments",               "DFM", "Diversified",       "Mid",   "HALAL",  ""),
    ("EIBANK.DU",     "Emirates Islamic Bank",           "DFM", "Banking-Islamic",   "Large", "HALAL",  ""),
    ("EMAAR.DU",      "Emaar Properties",                "DFM", "Real-Estate",       "Mega",  "HALAL",  ""),
    ("EMAARDEV.DU",   "Emaar Development",               "DFM", "Real-Estate",       "Large", "HALAL",  ""),
    ("EMIRATESREIT.DU","Emirates REIT",                  "DFM", "Real-Estate-REIT",  "Small", "HALAL",  ""),
    ("EMPOWER.DU",    "Empower",                         "DFM", "Utilities-Cooling", "Large", "HALAL",  ""),
    ("GFH.DU",        "GFH Financial Group",             "DFM", "Financial-Islamic", "Mid",   "HALAL",  ""),
    ("PARKIN.DU",     "Parkin Company",                  "DFM", "Industrials",       "Mid",   "HALAL",  ""),
    ("SALIK.DU",      "Salik (Dubai Tolls)",             "DFM", "Industrials-Tolls", "Large", "HALAL",  ""),
    ("SHUAA.DU",      "SHUAA Capital",                   "DFM", "Financial-Islamic", "Small", "VERIFY", "Islamic asset mgmt — ratio confirm"),
    ("SUKOON.DU",     "Sukoon Insurance",                "DFM", "Insurance-Takaful", "Mid",   "HALAL",  "Takaful (Shariah-compliant insurance)"),
    ("TABREED.DU",    "National Central Cooling Co",     "DFM", "Utilities-Cooling", "Mid",   "HALAL",  ""),
    ("TALABAT.DU",    "Talabat Holding",                 "DFM", "Tech-Consumer",     "Large", "HALAL",  ""),
    ("TECOM.DU",      "TECOM Group",                     "DFM", "Real-Estate",       "Mid",   "HALAL",  ""),
    ("UPP.DU",        "Union Properties",                "DFM", "Real-Estate",       "Small", "VERIFY", "Debt ratio historically borderline"),

    # --- Less liquid / verify ---
    ("ABNIC.AD",      "Al Buhaira National Insurance",   "ADX", "Insurance",         "Small", "VERIFY", "Confirm if takaful vs conventional"),
    ("MANAZEL.AD",    "Manazel",                         "ADX", "Real-Estate",       "Small", "VERIFY", "Recent restructuring — confirm status"),
    ("NCC.AD",        "Nat'l Corp Tourism & Hotels",     "ADX", "Consumer-Discretionary","Mid","VERIFY","Hotel revenue mix (alcohol exposure)"),
    ("NGI.DU",        "National General Insurance",      "DFM", "Insurance",         "Small", "VERIFY", "Confirm takaful classification"),
    ("ORASCONST.DU",  "Orascom Construction",            "DFM", "Construction",      "Mid",   "VERIFY", "Confirm DFM listing + ratios"),
]

EXCLUDED = [
    # Conventional banks — interest-based business model fails AAOIFI screen
    ("ADCB.AD",       "Abu Dhabi Commercial Bank",       "ADX", "Banking-Conv",      "Mega",  "EXCLUDE_CONV", "Conventional bank"),
    ("FAB.AD",        "First Abu Dhabi Bank",            "ADX", "Banking-Conv",      "Mega",  "EXCLUDE_CONV", "Conventional bank (merger of NBAD + FGB)"),
    ("CBI.AD",        "Commercial Bank International",   "ADX", "Banking-Conv",      "Mid",   "EXCLUDE_CONV", "Conventional bank"),
    ("RAKBANK.AD",    "RAK Bank",                        "ADX", "Banking-Conv",      "Mid",   "EXCLUDE_CONV", "Conventional bank"),
    ("INVESTBANK.AD", "Invest Bank",                     "ADX", "Banking-Conv",      "Small", "EXCLUDE_CONV", "Conventional bank"),
    ("NBF.AD",        "National Bank of Fujairah",       "ADX", "Banking-Conv",      "Mid",   "EXCLUDE_CONV", "Conventional bank"),
    ("UAB.AD",        "United Arab Bank",                "ADX", "Banking-Conv",      "Small", "EXCLUDE_CONV", "Conventional bank"),
    ("BOS.AD",        "Bank of Sharjah",                 "ADX", "Banking-Conv",      "Mid",   "EXCLUDE_CONV", "Conventional bank"),
    ("ENBD.DU",       "Emirates NBD",                    "DFM", "Banking-Conv",      "Mega",  "EXCLUDE_CONV", "Conventional bank"),
    ("CBD.DU",        "Commercial Bank of Dubai",        "DFM", "Banking-Conv",      "Large", "EXCLUDE_CONV", "Conventional bank"),
    ("MASHREQ.DU",    "Mashreq Bank",                    "DFM", "Banking-Conv",      "Large", "EXCLUDE_CONV", "Conventional bank"),
]


def _style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="left", vertical="center")
    side = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 22


def _autosize(ws, headers, rows):
    for i, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for r in rows:
            v = "" if i - 1 >= len(r) or r[i - 1] is None else str(r[i - 1])
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)


def _write_sheet(ws, headers, rows, status_col_index=None):
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    _style_header(ws, len(headers))
    _autosize(ws, headers, rows)
    # Status colouring
    if status_col_index is not None:
        for i in range(2, len(rows) + 2):
            cell = ws.cell(row=i, column=status_col_index)
            v = (cell.value or "").upper()
            if v == "HALAL":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(color="006100", bold=True)
            elif v == "VERIFY":
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
                cell.font = Font(color="9C5700", bold=True)
            elif v == "EXCLUDE_CONV":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)
    ws.freeze_panes = "A2"


def build():
    wb = Workbook()
    headers = ["Ticker", "Name", "Exchange", "Sector", "Cap Tier",
               "Halal Status", "Notes"]

    ws1 = wb.active
    ws1.title = "Halal Universe"
    _write_sheet(ws1, headers, HALAL, status_col_index=6)

    ws2 = wb.create_sheet("Excluded")
    _write_sheet(ws2, headers, EXCLUDED, status_col_index=6)

    ws3 = wb.create_sheet("About")
    about = [
        ["Field", "Value"],
        ["File", "uae_stocks.xlsx"],
        ["Author", "AIG CEO (autonomous build, Session 1, 2026-05-20)"],
        ["Status", "WORKING PROXY — pending Ahmed's authoritative ADIB UAE list"],
        ["Methodology", "AAOIFI sector + financial ratio screen on publicly-known ADX + DFM listings"],
        ["Excluded", "Conventional banks (interest-based) — see 'Excluded' sheet"],
        ["Verify",   "Names where sector or revenue mix is borderline (insurance, hotels, investment cos) — Ahmed to confirm"],
        ["Engine input", "Derived plain-text list at universe/uae_tickers.txt — engine reads this"],
        ["Override", "Drop authoritative list at universe/uae_tickers.txt to replace"],
        ["Rule 15", "Long-only — strategy enforces; universe-side this is a no-op"],
        ["Rule 16", "No leverage — strategy enforces"],
        ["Tickers (yfinance)", "ADX: <symbol>.AD; DFM: <symbol>.DU"],
        ["Note", "Some UAE tickers may not be available on yfinance — engine will skip with 'data error' and log in audit_trail.md"],
    ]
    for row in about:
        ws3.append(row)
    _style_header(ws3, 2)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 90
    ws3.freeze_panes = "A2"

    wb.save(XLSX_PATH)

    # Derived txt for engine: HALAL only (skip VERIFY/EXCLUDE)
    active = [t[0] for t in HALAL if t[5].upper() == "HALAL"]
    verify = [t[0] for t in HALAL if t[5].upper() == "VERIFY"]
    with open(TXT_PATH, "w", encoding="utf-8") as fh:
        fh.write("# UAE Shariah-compliant universe (CEO-curated proxy, 2026-05-20)\n")
        fh.write("# Source: universe/uae_stocks.xlsx — 'Halal Universe' sheet, status=HALAL.\n")
        fh.write("# Drop Ahmed's authoritative ADIB UAE list over this file to override.\n")
        fh.write(f"# {len(active)} HALAL tickers (engine consumes these).\n")
        fh.write(f"# {len(verify)} VERIFY tickers commented out below — uncomment after Ahmed confirms.\n")
        fh.write("# ADX suffix .AD ; DFM suffix .DU\n\n")
        for t in active:
            fh.write(f"{t}\n")
        fh.write("\n# --- VERIFY (commented; need Ahmed's confirm) ---\n")
        for t in verify:
            fh.write(f"# {t}\n")

    print(f"Wrote {XLSX_PATH}")
    print(f"Wrote {TXT_PATH}")
    print(f"Active HALAL: {len(active)}  | VERIFY pending: {len(verify)}  "
          f"| EXCLUDED: {len(EXCLUDED)}")


if __name__ == "__main__":
    build()
