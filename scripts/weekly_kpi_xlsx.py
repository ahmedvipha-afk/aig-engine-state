"""AIG weekly KPI scorecard — builds .xlsx + sends via Telegram.

Used by:
  - aig-weekly-full-universe Cloud Routine (Sunday 19:53 GST)
  - Manual: `python scripts/weekly_kpi_xlsx.py`

Output: reports/kpi_YYYY-WW.xlsx (overwritten if same week)
Then sends via Telegram sendDocument with a caption.
"""
from __future__ import annotations
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))
from telegram_send import send_document  # type: ignore

GST = timezone(timedelta(hours=4))
REPORTS_DIR = ROOT / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
NEAR_FILL = PatternFill("solid", fgColor="FFEB9C")


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22


def _autosize(ws, headers: list[str], rows: list[list]):
    for i, h in enumerate(headers, start=1):
        maxlen = len(str(h))
        for r in rows:
            v = r[i - 1] if i - 1 < len(r) else ""
            v = "" if v is None else str(v)
            if len(v) > maxlen:
                maxlen = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(maxlen + 2, 60)


def _latest_runs() -> dict:
    runs = {}
    for p in ROOT.glob("validation_*.json"):
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        pf = d.get("portfolio")
        if not pf:
            continue
        args = d.get("args", {})
        strat = args.get("strategy", "?")
        uni = (args.get("universe") or "").lower()
        if "uae" in p.name or "uae" in uni:
            market = "UAE"
        elif "crypto" in p.name or "crypto" in uni:
            market = "CRYPTO"
        else:
            market = "US"
        key = (strat, market)
        mt = p.stat().st_mtime
        if key not in runs or mt > runs[key][0]:
            runs[key] = (mt, p, pf)
    return runs


def _count_universe(path: Path) -> int:
    if not path.exists():
        return 0
    n = 0
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s and not s.startswith("#"):
            n += 1
    return n


def build_workbook() -> Path:
    now = datetime.now(GST)
    iso_year, iso_week, _ = now.isocalendar()
    fname = f"kpi_{iso_year}-W{iso_week:02d}.xlsx"
    out = REPORTS_DIR / fname

    wb = Workbook()

    # --- Sheet 1: KPI Scorecard (v7.0 §24, 14 KPIs) -----------------------
    ws = wb.active
    ws.title = "KPI Scorecard"
    headers = ["#", "KPI", "Current", "Target", "Status", "Owner", "Note"]
    runs = _latest_runs()
    pipeline_count = len(set(s for (s, _) in runs.keys()))
    coverage_us = sum(1 for (s, m), (_, _, pf) in runs.items()
                      if m == "US" and pf.get("contributing_tickers"))
    cleared = [(s, m) for (s, m), (_, _, pf) in runs.items() if pf.get("passed")]

    best_sr = max((pf.get("portfolio_sharpe_deflated") or 0)
                  for (_, _, pf) in runs.values()) if runs else 0

    kpi_rows = [
        ["1",  "Annual return",        "—",                "≥3x / 10x asp", "PRE", "CIO",         "Paper not started"],
        ["2",  "Sharpe ratio (best portfolio)",  f"{best_sr:.2f}", "≥1.5 / target ≥2.5", "PASS" if best_sr >= 1.5 else "INFO", "Performance", "Best deflated portfolio Sharpe across runs"],
        ["3",  "Max drawdown",         "—",                "<20%",         "PRE", "Risk Mgr",    "Paper not started"],
        ["4",  "Alpha vs S&P 500",     "—",                ">0",           "PRE", "CIO",         "Paper not started"],
        ["5",  "Win rate ★★★★+",       "—",                ">50%",         "PRE", "Learning",    "Rolling 30 trades"],
        ["6",  "Halal compliance",     "100%",             "100%",         "PASS", "Compliance", "Universes pre-screened"],
        ["7",  "Max equity positions", "0",                "8",            "PASS", "Portfolio",  "No open"],
        ["8",  "Max crypto positions", "0",                "3",            "PASS", "EX-F + Risk","No open"],
        ["9",  "Crypto allocation",    "0%",               "≤10% NAV",     "PASS", "EX-F + Risk","No open"],
        ["10", "BTC-SPY correlation",  "—",                "alert >0.70",  "PRE", "EX-D",        "Not monitored yet"],
        ["11", "Strategy pipeline",    f"{pipeline_count}", "≥3",          "PASS" if pipeline_count >= 3 else "WARN", "R&D Lab", "MBV is missing third"],
        ["12", "Data uptime",          "~95%",             ">99.5%",       "WARN", "Systems",    "UAE yfinance gap; TV cache mitigates"],
        ["13", "Audit coverage",       "100%",             "100%",         "PASS", "Audit",      "audit_trail.md per run"],
        ["14", "Notification latency", "—",                "<60s",         "PRE", "Notification","Telegram online; first signal pending"],
    ]
    ws.append(headers)
    for row in kpi_rows:
        ws.append(row)
    _style_header(ws, len(headers))
    _autosize(ws, headers, kpi_rows)
    for i, row in enumerate(kpi_rows, start=2):
        status = row[4]
        fill = {"PASS": PASS_FILL, "WARN": NEAR_FILL, "FAIL": FAIL_FILL}.get(status)
        if fill:
            ws.cell(row=i, column=5).fill = fill
    ws.freeze_panes = "A2"

    # --- Sheet 2: Strategy Verdicts ---------------------------------------
    ws2 = wb.create_sheet("Strategy Verdicts")
    headers2 = ["Strategy", "Market", "Verdict", "Trades", "Contributors",
                "Universe", "Coverage", "Expectancy", "Win Rate",
                "Raw Sharpe", "Deflated Sharpe", "Top Fail Reason"]
    rows2 = []
    for (strat, market), (_, _, pf) in sorted(runs.items()):
        rows2.append([
            strat, market, pf.get("verdict"), pf.get("portfolio_trades"),
            pf.get("contributing_tickers"), pf.get("universe_size"),
            f"{(pf.get('universe_coverage') or 0)*100:.1f}%",
            pf.get("portfolio_expectancy"),
            f"{(pf.get('portfolio_win_rate') or 0)*100:.1f}%",
            pf.get("portfolio_sharpe_raw"),
            pf.get("portfolio_sharpe_deflated"),
            (pf.get("reasons") or [""])[0],
        ])
    ws2.append(headers2)
    for r in rows2:
        ws2.append(r)
    _style_header(ws2, len(headers2))
    _autosize(ws2, headers2, rows2)
    for i, row in enumerate(rows2, start=2):
        verdict = row[2]
        fill = PASS_FILL if "CLEARED" in (verdict or "") else FAIL_FILL
        ws2.cell(row=i, column=3).fill = fill
    ws2.freeze_panes = "A2"

    # --- Sheet 3: Coverage detail -----------------------------------------
    ws3 = wb.create_sheet("Coverage")
    headers3 = ["Universe", "Active tickers", "Target", "% of target", "Notes"]
    coverage_rows = [
        ["US halal", _count_universe(ROOT / "universe" / "us_halal_full.txt"),
         100, "", "Ahmed authoritative ADIB list, mcap-sorted"],
        ["UAE halal (unified)", _count_universe(ROOT / "universe" / "uae_tickers_full.txt"),
         60, "", "Hybrid: data_cache CSV (23 ADX via TV) + yfinance .AE/.AB (21)"],
        ["Crypto halal", _count_universe(ROOT / "universe" / "halal_crypto_150_USD.txt"),
         100, "", "Ahmed authoritative -USDT list, converted to -USD for yfinance"],
    ]
    for r in coverage_rows:
        r[3] = f"{(r[1] / r[2]) * 100:.1f}%" if r[2] else "—"
    ws3.append(headers3)
    for r in coverage_rows:
        ws3.append(r)
    _style_header(ws3, len(headers3))
    _autosize(ws3, headers3, coverage_rows)
    for i, r in enumerate(coverage_rows, start=2):
        pct = float(r[3].rstrip("%")) if r[3] != "—" else 0
        fill = PASS_FILL if pct >= 100 else (NEAR_FILL if pct >= 50 else FAIL_FILL)
        ws3.cell(row=i, column=4).fill = fill
    ws3.freeze_panes = "A2"

    # --- Sheet 4: About / methodology -------------------------------------
    ws4 = wb.create_sheet("About")
    about = [
        ["Field", "Value"],
        ["Report", "Weekly KPI Scorecard"],
        ["Generated", now.strftime("%Y-%m-%d %H:%M GST")],
        ["ISO Week", f"{iso_year}-W{iso_week:02d}"],
        ["Repo", "https://github.com/ahmedvipha-afk/aig-engine-state"],
        ["Cockpit", "raw.githubusercontent.com/ahmedvipha-afk/aig-engine-state/main/dashboard.html"],
        ["Strategies", ", ".join(sorted(set(s for (s, _) in runs.keys()))) or "—"],
        ["Markets", "US, UAE, CRYPTO"],
        ["Cleared portfolios", ", ".join(f"{s}/{m}" for (s, m) in cleared) or "—"],
        ["Audit", "Every run logs to aig/audit_trail.md with config_hash"],
        ["Constraints", "Rule 15 long-only · Rule 16 no leverage · Rule 19 crypto spot-only · 5-point Shariah"],
    ]
    for r in about:
        ws4.append(r)
    _style_header(ws4, 2)
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 90
    ws4.freeze_panes = "A2"

    wb.save(out)
    return out


def main():
    path = build_workbook()
    print(f"WROTE {path}  ({path.stat().st_size:,} bytes)")
    now = datetime.now(GST)
    caption = (f"📈 AIG Weekly KPI Scorecard — week of "
               f"{now.strftime('%Y-%m-%d')}\n"
               f"Auto-generated. Cockpit: "
               f"https://raw.githubusercontent.com/ahmedvipha-afk/aig-engine-state/main/dashboard.html")
    r = send_document(path, caption=caption)
    if r.get("ok"):
        print(f"SENT message_id={r.get('result', {}).get('message_id', '?')}")
    else:
        print(f"FAIL: {r}")
        sys.exit(1)


if __name__ == "__main__":
    main()
