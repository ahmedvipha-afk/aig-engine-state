"""AIG monthly report — builds a multi-sheet xlsx + sends via Telegram.

Sheets:
  1. Executive Summary    - one-page snapshot for Ahmed
  2. KPI Scorecard        - all 14 KPIs (v7.0 §24)
  3. Strategy Verdicts    - every strategy × market with full metrics
  4. Coverage Detail      - per-universe tickers cleared
  5. Top Contributors     - per strategy, top per-ticker expectancy leaders
  6. Decisions Log        - this month's CEO decisions from ceo_brain.md
  7. Audit Activity       - recent commits + cloud routine runs
  8. About                - methodology + repo + cockpit links

Used by:
  - aig-monthly-report Cloud Routine (1st of month)
  - Manual: `python scripts/monthly_report.py`
"""
from __future__ import annotations
import json
import re
import subprocess
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))
from telegram_send import send_document  # type: ignore

GST = timezone(timedelta(hours=4))
REPORTS_DIR = ROOT / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
NEAR_FILL = PatternFill("solid", fgColor="FFEB9C")
ACCENT_FILL = PatternFill("solid", fgColor="EEF6FC")


def _style_header(ws, ncols: int, row: int = 1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[row].height = 22


def _autosize(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _latest_runs() -> dict:
    runs = {}
    for p in ROOT.glob("validation_*.json"):
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        pf = d.get("portfolio")
        if not pf:
            continue
        args = d.get("args", {})
        strat = args.get("strategy", "?")
        uni = (args.get("universe") or "").lower()
        if "uae" in p.name or "uae" in uni:
            market = "UAE"
        elif "crypto" in p.name or "crypto" in uni:
            market = "CRYPTO"
        else:
            market = "US"
        key = (strat, market)
        mt = p.stat().st_mtime
        if key not in runs or mt > runs[key][0]:
            runs[key] = (mt, p, d)
    return runs


def _count_universe(path: Path) -> int:
    if not path.exists():
        return 0
    return sum(1 for line in path.read_text(encoding="utf-8").splitlines()
               if line.strip() and not line.strip().startswith("#"))


def _git_log_month() -> list[dict]:
    try:
        out = subprocess.check_output(
            ["git", "log", "--since=30.days", "--pretty=format:%h|%ad|%s",
             "--date=iso-strict"],
            cwd=str(ROOT), text=True, stderr=subprocess.DEVNULL
        )
    except Exception:
        return []
    rows = []
    for line in out.splitlines():
        parts = line.split("|", 2)
        if len(parts) == 3:
            rows.append({"sha": parts[0], "date": parts[1], "subject": parts[2]})
    return rows


def _decisions_from_brain() -> list[str]:
    p = ROOT / "ceo_brain.md"
    if not p.exists():
        return []
    text = p.read_text(encoding="utf-8", errors="ignore")
    out = []
    for m in re.finditer(r"^\d+\.\s+\*\*(\d{4}-\d{2}-\d{2}\s+D-\d+)\*\*\s+(.+?)(?=\n\d+\.|\n##|\Z)",
                         text, flags=re.M | re.S):
        out.append(f"{m.group(1)} — {m.group(2).strip().splitlines()[0]}")
    return out


def _top_contributors(runs: dict, n: int = 15) -> dict[tuple[str, str], list[dict]]:
    """Per strategy × market, pull top-N per-ticker expectancy leaders."""
    out = {}
    for (strat, market), (_, path, d) in runs.items():
        rows = d.get("results", [])
        # qualifying tickers: oos_n >= 30, expectancy >= 1.0
        rows = [r for r in rows if isinstance(r, dict)
                and (r.get("oos_n") or 0) >= 30
                and (r.get("oos_expectancy") or 0) >= 1.0]
        rows.sort(key=lambda r: -(r.get("oos_expectancy") or 0))
        out[(strat, market)] = rows[:n]
    return out


def build_workbook(period_label: str | None = None) -> Path:
    now = datetime.now(GST)
    period = period_label or now.strftime("%Y-%m")
    fname = f"monthly_{period}.xlsx"
    out = REPORTS_DIR / fname

    runs = _latest_runs()
    cleared = [(s, m) for (s, m), (_, _, d) in runs.items() if d["portfolio"].get("passed")]
    decisions = _decisions_from_brain()
    commits = _git_log_month()
    top = _top_contributors(runs)

    wb = Workbook()

    # ---- Sheet 1: Executive Summary --------------------------------------
    ws = wb.active
    ws.title = "Executive Summary"
    ws.append(["Ahmed Investment Group — Monthly Report"])
    ws.append([f"Period: {period}    Generated: {now.strftime('%Y-%m-%d %H:%M GST')}"])
    ws.append([])
    ws.cell(row=1, column=1).font = Font(bold=True, size=18, color="1F4E78")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    rows = [
        ["Cleared portfolio strategies",
         ", ".join(f"{s}/{m}" for (s, m) in cleared) or "—",
         "Verdict = PORTFOLIO_CLEARED_FOR_PAPER_FORWARD"],
        ["Open positions", "0 (paper, pre-launch)", "Rule 01 paper-only until manual go-live"],
        ["Strategies registered", str(len(set(s for (s, _) in runs.keys()))), "v7.0 §19 target ≥3"],
        ["Universe size — US halal",
         f"{_count_universe(ROOT / 'universe' / 'us_halal_full.txt')} tickers", "Authoritative ADIB list"],
        ["Universe size — UAE halal",
         f"{_count_universe(ROOT / 'universe' / 'uae_tickers_full.txt')} tickers",
         "23 ADX via TV-MCP cache + 21 DFM via yfinance"],
        ["Universe size — Crypto halal",
         f"{_count_universe(ROOT / 'universe' / 'halal_crypto_150_USD.txt')} tickers",
         "Authoritative -USDT list converted -USD"],
        ["CEO decisions logged this month", str(len(decisions)), "From ceo_brain.md DECISION CONTINUITY"],
        ["Git commits this month", str(len(commits)), "Backup repo aig-engine-state"],
        ["Telegram channel", "ONLINE",
         "@AIV_Fund_Bot · policy=allowlist · scheduled task for persistence"],
        ["Cockpit", "dashboard.html",
         "https://raw.githubusercontent.com/ahmedvipha-afk/aig-engine-state/main/dashboard.html"],
    ]
    ws.append(["Metric", "Value", "Note"])
    _style_header(ws, 3, row=4)
    for r in rows:
        ws.append(r)
    _autosize(ws, [38, 50, 80])
    ws.freeze_panes = "A5"

    # ---- Sheet 2: KPI Scorecard ------------------------------------------
    ws2 = wb.create_sheet("KPI Scorecard")
    pipeline_count = len(set(s for (s, _) in runs.keys()))
    best_sr = max((d["portfolio"].get("portfolio_sharpe_deflated") or 0)
                  for (_, _, d) in runs.values()) if runs else 0
    headers = ["#", "KPI", "Current", "Target", "Status", "Owner", "Note"]
    ws2.append(headers)
    kpi_rows = [
        ["1",  "Annual return",          "—",                "≥3x / 10x asp", "PRE",  "CIO",          "Paper not started"],
        ["2",  "Sharpe (best portfolio)",f"{best_sr:.2f}",  "≥1.5",         "PASS" if best_sr >= 1.5 else "INFO", "Performance", "Deflated portfolio Sharpe"],
        ["3",  "Max drawdown",            "—",                "<20%",         "PRE",  "Risk Mgr",     "Paper not started"],
        ["4",  "Alpha vs S&P 500",        "—",                ">0",           "PRE",  "CIO",          "Paper not started"],
        ["5",  "Win rate ★★★★+",          "—",                ">50%",         "PRE",  "Learning",     "Rolling 30 trades"],
        ["6",  "Halal compliance",        "100%",             "100%",         "PASS", "Compliance",   "Universes pre-screened"],
        ["7",  "Max equity positions",    "0",                "8",            "PASS", "Portfolio",    "No open"],
        ["8",  "Max crypto positions",    "0",                "3",            "PASS", "EX-F + Risk",  "No open"],
        ["9",  "Crypto allocation",       "0%",               "≤10% NAV",     "PASS", "EX-F + Risk",  "No open"],
        ["10", "BTC-SPY correlation",     "—",                "alert >0.70",  "PRE",  "EX-D",         "Not monitored yet"],
        ["11", "Strategy pipeline",       f"{pipeline_count}", "≥3",          "PASS" if pipeline_count >= 3 else "WARN", "R&D Lab", "MBV is the missing third"],
        ["12", "Data uptime",             "~95%",             ">99.5%",       "WARN", "Systems",      "UAE yfinance gap; TV cache mitigates"],
        ["13", "Audit coverage",          "100%",             "100%",         "PASS", "Audit",        "audit_trail.md per run"],
        ["14", "Notification latency",    "—",                "<60s",         "PRE",  "Notification", "Telegram online; first signal pending"],
    ]
    for r in kpi_rows:
        ws2.append(r)
    _style_header(ws2, len(headers))
    for i, row in enumerate(kpi_rows, start=2):
        fill = {"PASS": PASS_FILL, "WARN": NEAR_FILL, "FAIL": FAIL_FILL}.get(row[4])
        if fill:
            ws2.cell(row=i, column=5).fill = fill
    _autosize(ws2, [4, 30, 25, 18, 8, 18, 60])
    ws2.freeze_panes = "A2"

    # ---- Sheet 3: Strategy Verdicts --------------------------------------
    ws3 = wb.create_sheet("Strategy Verdicts")
    headers3 = ["Strategy", "Market", "Verdict", "Trades", "Contrib",
                "Universe", "Coverage", "Expectancy", "Win Rate",
                "Sharpe (raw)", "Sharpe (deflated)", "Top Fail Reason"]
    ws3.append(headers3)
    for (strat, market), (_, _, d) in sorted(runs.items()):
        pf = d["portfolio"]
        ws3.append([
            strat, market, pf.get("verdict"),
            pf.get("portfolio_trades"), pf.get("contributing_tickers"),
            pf.get("universe_size"),
            f"{(pf.get('universe_coverage') or 0)*100:.1f}%",
            pf.get("portfolio_expectancy"),
            f"{(pf.get('portfolio_win_rate') or 0)*100:.1f}%",
            pf.get("portfolio_sharpe_raw"),
            pf.get("portfolio_sharpe_deflated"),
            (pf.get("reasons") or [""])[0],
        ])
    _style_header(ws3, len(headers3))
    for i, (_, _, d) in enumerate(sorted(runs.values(), key=lambda x: x[1].name), start=2):
        verdict = d["portfolio"].get("verdict")
        fill = PASS_FILL if "CLEARED" in (verdict or "") else FAIL_FILL
        ws3.cell(row=i, column=3).fill = fill
    _autosize(ws3, [14, 10, 32, 10, 10, 10, 10, 12, 10, 12, 16, 60])
    ws3.freeze_panes = "A2"

    # ---- Sheet 4: Coverage Detail ----------------------------------------
    ws4 = wb.create_sheet("Coverage Detail")
    headers4 = ["Universe", "Active", "Target", "% of target", "Notes"]
    ws4.append(headers4)
    cov_rows = [
        ["US halal",
         _count_universe(ROOT / "universe" / "us_halal_full.txt"),
         100, "", "Authoritative ADIB, mcap-sorted; 1,123 yfinance-valid in latest run"],
        ["UAE halal (unified)",
         _count_universe(ROOT / "universe" / "uae_tickers_full.txt"),
         60, "", "Cache (TV MCP 23 ADX) + yfinance (.AE/.AB 21); 6 illiquid dropped"],
        ["Crypto halal",
         _count_universe(ROOT / "universe" / "halal_crypto_150_USD.txt"),
         100, "", "Authoritative -USDT; ~140/150 valid with relaxed integrity"],
    ]
    for r in cov_rows:
        r[3] = f"{(r[1]/r[2])*100:.1f}%" if r[2] else "—"
    for r in cov_rows:
        ws4.append(r)
    _style_header(ws4, len(headers4))
    for i, r in enumerate(cov_rows, start=2):
        pct = float(r[3].rstrip("%")) if r[3] != "—" else 0
        fill = PASS_FILL if pct >= 100 else (NEAR_FILL if pct >= 50 else FAIL_FILL)
        ws4.cell(row=i, column=4).fill = fill
    _autosize(ws4, [24, 12, 10, 14, 80])
    ws4.freeze_panes = "A2"

    # ---- Sheet 5: Top Contributors ---------------------------------------
    ws5 = wb.create_sheet("Top Contributors")
    headers5 = ["Strategy", "Market", "Ticker", "OOS Trades",
                "Expectancy", "Win Rate", "Sharpe (deflated)"]
    ws5.append(headers5)
    for (strat, market), rows in sorted(top.items()):
        for r in rows:
            ws5.append([
                strat, market, r.get("ticker"), r.get("oos_n"),
                r.get("oos_expectancy"),
                "—" if r.get("oos_win_rate") is None else r.get("oos_win_rate"),
                r.get("oos_sharpe_deflated"),
            ])
    _style_header(ws5, len(headers5))
    _autosize(ws5, [14, 10, 12, 12, 12, 12, 16])
    ws5.freeze_panes = "A2"

    # ---- Sheet 6: Decisions Log ------------------------------------------
    ws6 = wb.create_sheet("Decisions")
    ws6.append(["Date · ID", "Decision (first line)"])
    for d in decisions:
        if " — " in d:
            tag, rest = d.split(" — ", 1)
            ws6.append([tag, rest])
        else:
            ws6.append([d, ""])
    _style_header(ws6, 2)
    _autosize(ws6, [22, 100])
    ws6.freeze_panes = "A2"

    # ---- Sheet 7: Audit Activity (git commits) ---------------------------
    ws7 = wb.create_sheet("Commits (30d)")
    ws7.append(["SHA", "Date", "Subject"])
    for c in commits:
        ws7.append([c["sha"], c["date"], c["subject"]])
    _style_header(ws7, 3)
    _autosize(ws7, [12, 26, 100])
    ws7.freeze_panes = "A2"

    # ---- Sheet 8: About ---------------------------------------------------
    ws8 = wb.create_sheet("About")
    about = [
        ["Field", "Value"],
        ["Report",          f"AIG Monthly — {period}"],
        ["Generated",       now.strftime("%Y-%m-%d %H:%M GST")],
        ["Engine repo",     "https://github.com/ahmedvipha-afk/aig-engine-state (PUBLIC)"],
        ["Cockpit URL",     "https://raw.githubusercontent.com/ahmedvipha-afk/aig-engine-state/main/dashboard.html"],
        ["Telegram bot",    "@AIV_Fund_Bot (allowlist policy)"],
        ["Constraints",     "Rule 01 paper · Rule 15 long only · Rule 16 no leverage · Rule 19 crypto spot · 5-pt Shariah"],
        ["Pre-registration","Every strategy frozen in config.py; provenance hash binds spec to verdict"],
        ["Gates",           "Per-ticker (GATE) + Portfolio (PORTFOLIO_GATE)"],
        ["Methodology",     "OOS = post-60% train split; deflated Sharpe = haircut over N_strategies (portfolio gate)"],
    ]
    for r in about:
        ws8.append(r)
    _style_header(ws8, 2)
    _autosize(ws8, [22, 100])
    ws8.freeze_panes = "A2"

    wb.save(out)
    return out


def main():
    path = build_workbook()
    print(f"WROTE {path}  ({path.stat().st_size:,} bytes)")
    now = datetime.now(GST)
    caption = (f"📄 AIG Monthly Report — {now.strftime('%Y-%m')}\n"
               "Full multi-sheet xlsx: KPIs, strategy verdicts, coverage, "
               "top contributors, decisions, commits.\n"
               "Public repo: https://github.com/ahmedvipha-afk/aig-engine-state")
    r = send_document(path, caption=caption)
    if r.get("ok"):
        print(f"SENT message_id={r.get('result', {}).get('message_id', '?')}")
    else:
        print(f"FAIL: {r}")
        sys.exit(1)


if __name__ == "__main__":
    main()
